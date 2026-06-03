"""
Run this once to mark M2 Done and add the change log entry.
  python scripts/update_tracksheet_m2.py
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path(__file__).resolve().parents[1]
PATH = REPO / "TRACKSHEET.xlsx"

def solid(h): return PatternFill("solid", fgColor=h)
def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.load_workbook(PATH)
ms = wb["Milestones"]
for row in ms.iter_rows():
    if row[0].value == "M2":
        row[5].value = "Done"
        row[6].value = "pipeline/cluster/ built: embed.py (sentence-transformers), group.py (cosine threshold + connected components, manual override), run.py (CLI). 2026-05-31"
        for cell in row:
            if cell.column != 3:
                cell.fill = solid("E2EFDA")
        break

cl = wb["Change Log"]
r = cl.max_row + 1
cl.row_dimensions[r].height = 36
fill = solid("F7F7F7") if r % 2 else solid("FFFFFF")
entry = (
    "2026-05-31", "M2 — Cluster",
    "Built pipeline/cluster/embed.py (sentence-transformers all-MiniLM-L6-v2, L2-normalised), "
    "group.py (threshold-based cosine similarity + connected components; manual override), "
    "run.py (CLI: --source golden/live, --threshold, --manual). "
    "Outputs cluster JSON to data/events/<beat>/.",
    "M2 definition of done met — auto-cluster + manual override ready; correctness check runs against golden event_001",
    "Claude (Cowork)",
)
for col, val in enumerate(entry, 1):
    c = cl.cell(row=r, column=col, value=val)
    c.font = Font(name="Arial", size=9)
    c.fill = fill
    c.border = thin()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb.save(PATH)
print(f"Saved: {PATH}")

# Verify
wb2 = openpyxl.load_workbook(PATH, data_only=True)
for row in wb2["Milestones"].iter_rows(values_only=True):
    if row[0] in ("M0","M1","M2","M3"):
        print(f"  {row[0]}: {row[5]}")
